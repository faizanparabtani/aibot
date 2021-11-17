from flask import Flask, render_template, request
from chatterbot import ChatBot
from chatterbot.trainers import ChatterBotCorpusTrainer
import os
import openpyxl
from openpyxl.utils import get_column_letter

from chatterbot import ChatBot
from chatterbot.trainers import ListTrainer

from flask import jsonify, json

from opencage.geocoder import OpenCageGeocode



filenumber=int(os.listdir('saved_conversations')[-1])
filenumber=filenumber+1
file= open('saved_conversations/'+str(filenumber),"w+")
file.write('bot : Hi There! I am a medical chatbot. You can begin conversation by typing in a message and pressing enter.\n')
file.close()

app = Flask(__name__)


english_bot = ChatBot('Bot',
             storage_adapter='chatterbot.storage.SQLStorageAdapter',
             logic_adapters=[
   {
       'import_path': 'chatterbot.logic.BestMatch'
   },

],
trainer='chatterbot.trainers.ListTrainer')
english_bot.set_trainer(ListTrainer)

def get_doctor_from_city(diagnosis, city):
    filtered_doctors = []
    for i in diagnosis:
        if city in i[1]:
            filtered_doctors.append(i)
    return filtered_doctors


def get_doctor_from_diagnosis(diagnosis):
    wb = openpyxl.load_workbook('doctor_data.xlsx')
    ws = wb.active
    print(ws)
    doctor_list = []
    for row in range(2, ws.max_row+1):
        col = "M"
        if diagnosis == ws[col+str(row)].value:
            doctor_name = ws["A"+str(row)].value + " " + ws["B"+str(row)].value
            doctor_address = str(ws["D"+str(row)].value) + " " + str(ws["F"+str(row)].value) + " " + str(ws["G"+str(row)].value) + " " + str(ws["H"+str(row)].value)
            doctor_specialty = str(ws["L"+str(row)].value)
            doctor_contact = str(ws["L"+str(row)].value)
            doctor_list.append((doctor_name, doctor_address, doctor_specialty, doctor_contact))
    return doctor_list

def get_location():
    # Location
    key = '3265fdccd6a348a8b506da1cc60cd23e'
    geocoder = OpenCageGeocode(key)

    # Send location from geolocation to method below
    results = geocoder.reverse_geocode(19.0485642, 72.8240584)
    return results




diagnosis = ''
city = ''
doctor_response = None

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/get")
def get_bot_response():
    userText = request.args.get('msg')
    response = str(english_bot.get_response(userText))
    # Diagnosis
    user_response_list = userText.split(' ')
    bot_response = response.split(' ')

    global diagnosis
    global city
    global doctor_response

    if doctor_response != None:
        if 'Book an appointment' in userText or 'book an appointment' in userText or  'Show me available doctors ' in userText or  'show me available doctors ' in userText:
            print(doctor_response)
            return doctor_response

    if 'live' in user_response_list or 'stay' in user_response_list or 'reside' in user_response_list:
        city = user_response_list[-1]

    if 'fracture' in bot_response:
        diagnosis = 'Fracture'
    elif 'headache' in bot_response or 'headaches' in bot_response:
        diagnosis = 'Headache'
    elif 'cold.' in bot_response or 'cough.' in bot_response:
        diagnosis = 'Cold Cough'
    elif 'fever' in bot_response:
        diagnosis = 'Fever'
    elif 'anxiety.' in bot_response:
        diagnosis = 'Anxiety'
    elif 'itch' in bot_response or 'swallowing food' in bot_response or 'ear-wax' in bot_response:
        diagnosis = 'ENT'


    if diagnosis != "":
        get_diagnosis = get_doctor_from_diagnosis(diagnosis)
        if city != "":
            get_doctors = get_doctor_from_city(get_diagnosis, city)
            data = get_doctors
            doctor_response = app.response_class(
                response=json.dumps(data),
                mimetype='application/json'
            )


    appendfile=os.listdir('saved_conversations')[-1]
    appendfile= open('saved_conversations/'+str(filenumber),"a")
    appendfile.write('user : '+userText+'\n')
    appendfile.write('bot : '+response+'\n')
    appendfile.close()

    return response


if __name__ == "__main__":
    app.run()